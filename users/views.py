from django.contrib.auth import authenticate, login, logout
from django.shortcuts import render, redirect
from django.views.generic import TemplateView, ListView
from django.contrib.auth.mixins import LoginRequiredMixin
from rest_framework import generics, permissions
from rest_framework.exceptions import ValidationError, PermissionDenied
from rest_framework.response import Response
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.db.models import Count, Q
from .models import Order, OrderClosure, Accounting, FilterData, Users, ChatMessage
from notifications.models import Notification
from .serializers import OrderSerializer, OrderClosureSerializer, AccountingSerializer
from .forms import CustomLoginForm, OrderForm, FilterForm
from .permissions import IsAdmin, IsManager, IsDispecher, IsBugalter, IsFilter
from django.contrib import messages
from django.utils import timezone
from datetime import datetime, timedelta
import json
import logging
import pyttsx3
import threading
from django.core.paginator import Paginator
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import io

logger = logging.getLogger(__name__)


def speak_local(text):
    def run_speech():
        engine = pyttsx3.init()
        engine.setProperty('rate', 140)
        engine.setProperty('volume', 1.0)
        voices = engine.getProperty('voices')
        for voice in voices:
            if 'zira' in voice.name.lower() or 'female' in voice.name.lower():
                engine.setProperty('voice', voice.id)
                break
        engine.say(text)
        engine.runAndWait()

    thread = threading.Thread(target=run_speech)
    thread.start()


def custom_login_view(request):
    if request.method == 'POST':
        form = CustomLoginForm(request.POST)
        if form.is_valid():
            login_value = form.cleaned_data['login']
            password = form.cleaned_data['password']
            user = authenticate(request, login=login_value, password=password)
            if user is not None:
                login(request, user)
                logger.info(f"Foydalanuvchi {user.login} (ID: {user.id}) kirish qildi - {timezone.now()}")
                return redirect('/')
            else:
                messages.error(request, "Noto‘g‘ri login yoki parol")
                logger.warning(f"Kirish urunishi muvaffaqiyatsiz: {login_value} - {timezone.now()}")
    else:
        form = CustomLoginForm()
    return render(request, 'login.html', {'form': form})


def custom_logout_view(request):
    user = request.user
    logout(request)
    logger.info(f"Foydalanuvchi {user.login} (ID: {user.id}) chiqdi - {timezone.now()}")
    return redirect('custom_login')


def create_order_view(request):
    if request.user.status != 'manager':
        messages.error(request, "Faqat menejerlar buyurtma yaratishi mumkin.")
        logger.warning(
            f"Foydalanuvchi {request.user.login} (ID: {request.user.id}) buyurtma yaratishga ruxsatsiz urindi - {timezone.now()}")
        return redirect('/')
    if request.method == 'POST':
        form = OrderForm(request.POST)
        if form.is_valid():
            order = form.save(commit=False)
            order.manager = request.user
            order.save()
            logger.info(
                f"Buyurtma #{order.order_id} yaratildi - Menejer: {request.user.login} (ID: {request.user.id}) - {timezone.now()}")
            speak_local("New zayavka")
            messages.success(request, "Buyurtma muvaffaqiyatli yaratildi!")
            return redirect('/')
    else:
        form = OrderForm()
    return render(request, 'create_order.html', {'form': form})


def close_order_page(request, order_id):
    if request.user.status != 'dispecher':
        messages.error(request, "Faqat dispetcherlar buyurtmani yopishi mumkin.")
        logger.warning(
            f"Foydalanuvchi {request.user.login} (ID: {request.user.id}) buyurtma yopishga ruxsatsiz urindi - {timezone.now()}")
        return redirect('/')
    order = Order.objects.get(order_id=order_id)
    if request.method == 'POST':
        close_price = request.POST.get('close_price')
        price_turi = request.POST.get('price_turi')
        if not close_price or not price_turi:
            messages.error(request, "Iltimos, barcha maydonlarni to‘ldiring!")
            return render(request, 'close_order.html', {'order': order})
        OrderClosure.objects.create(
            order=order,
            dispecher=request.user,
            close_price=close_price,
            price_turi=price_turi
        )
        order.close_order_status = True
        order.save()
        logger.info(
            f"Buyurtma #{order.order_id} yopildi - Dispecher: {request.user.login} (ID: {request.user.id}) - {timezone.now()}")
        messages.success(request, "Buyurtma muvaffaqiyatli yopildi va filterchiga yuborildi!")
        return redirect('/')
    return render(request, 'close_order.html', {'order': order})


def filter_order_view(request, order_id):
    if request.user.status != 'filter':
        messages.error(request, "Faqat filtrchilar buyurtmani ko‘rishi mumkin.")
        logger.warning(
            f"Foydalanuvchi {request.user.login} (ID: {request.user.id}) filtr sahifasiga ruxsatsiz kirishga urindi - {timezone.now()}")
        return redirect('/')
    order = Order.objects.get(order_id=order_id)
    if not hasattr(order, 'closure'):
        messages.error(request, "Bu buyurtma hali yopilmagan!")
        return redirect('/filter/')
    if request.method == 'POST':
        form = FilterForm(request.POST)
        if form.is_valid():
            filter_data = form.save(commit=False)
            filter_data.order = order
            filter_data.dispecher_fullname = order.closure.dispecher.full_name
            filter_data.manager_fullname = order.manager.full_name
            filter_data.save()
            order.filter_status = True
            order.save()
            Accounting.objects.create(
                order=order,
                manager_price=order.price,
                dispecher_price=order.closure.close_price,
                payment_type=order.closure.price_turi,
                payment_status='jarayonda'
            )
            logger.info(
                f"Buyurtma #{order.order_id} filtrlandi - Filterchi: {request.user.login} (ID: {request.user.id}) - {timezone.now()}")
            messages.success(request, "Buyurtma bugalteriyaga yuborildi!")
            return redirect('/filter/')
    else:
        form = FilterForm()
    return render(request, 'filter_order.html', {'form': form, 'order': order})


def return_order_to_dispecher(request, order_id):
    if request.user.status != 'filter':
        messages.error(request, "Faqat filtrchilar buyurtmani qaytarishi mumkin.")
        logger.warning(
            f"Foydalanuvchi {request.user.login} (ID: {request.user.id}) buyurtma qaytarishga ruxsatsiz urindi - {timezone.now()}")
        return redirect('/filter/')
    order = Order.objects.get(order_id=order_id)
    if hasattr(order, 'closure'):
        order.closure.delete()
        order.close_order_status = False
        order.filter_status = False
        order.save()
        logger.info(
            f"Buyurtma #{order.order_id} dispecherga qaytarildi - Filterchi: {request.user.login} (ID: {request.user.id}) - {timezone.now()}")
        messages.success(request, "Buyurtma dispecherga qaytarildi!")
    else:
        messages.error(request, "Bu buyurtma yopilmagan!")
    return redirect('/filter/')


class OrderTemplateView(LoginRequiredMixin, TemplateView):
    template_name = 'index.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        print(f"User: {user}, Status: {user.status}")  # Debugging uchun
        if user.status == 'manager':
            orders = Order.objects.filter(manager=user)
        elif user.status == 'dispecher':
            orders = Order.objects.filter(close_order_status=False)
        elif user.status in ['admin', 'bugalter', 'filter']:
            orders = Order.objects.all()
        else:
            orders = Order.objects.none()

        # Filter turi (kun, oy, yil)
        filter_type = self.request.GET.get('filter_type', 'day')
        filter_date = self.request.GET.get('filter_date')
        if filter_date:
            selected_date = datetime.strptime(filter_date, '%Y-%m-%d')
            if filter_type == 'day':
                orders = orders.filter(created_at__date=selected_date)
            elif filter_type == 'month':
                start_date = selected_date.replace(day=1)
                end_date = (start_date + timedelta(days=31)).replace(day=1) - timedelta(seconds=1)
                orders = orders.filter(created_at__range=[start_date, end_date])
            elif filter_type == 'year':
                start_date = selected_date.replace(month=1, day=1)
                end_date = start_date.replace(month=12, day=31) + timedelta(seconds=-1)
                orders = orders.filter(created_at__range=[start_date, end_date])

        # Vaqtni hisoblash va rang aniqlash
        now = timezone.now()
        orders_with_time = []
        for order in orders:
            time_diff = now - order.created_at
            minutes = time_diff.total_seconds() / 60
            if minutes <= 30:
                time_color = 'green'
            elif minutes <= 90:
                time_color = 'yellow'
            else:
                time_color = 'red'
            orders_with_time.append({
                'order': order,
                'time_diff_minutes': minutes,
                'time_color': time_color
            })

        print(f"Orders: {orders}")  # Debugging uchun
        paginator = Paginator(orders_with_time, 10)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context['orders'] = page_obj
        context['page_obj'] = page_obj
        return context


class FilterTemplateView(LoginRequiredMixin, TemplateView):
    template_name = 'filter.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.status == 'filter':
            orders = Order.objects.filter(close_order_status=True, filter_status=False)
        else:
            orders = Order.objects.none()

        paginator = Paginator(orders, 10)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context['orders'] = page_obj
        context['page_obj'] = page_obj
        return context


@login_required
def undo_payment_view(request, order_id):
    if request.user.status != 'bugalter':
        messages.error(request, "Faqat bugalterlar to‘lovni bekor qilishi mumkin.")
        logger.warning(
            f"Foydalanuvchi {request.user.login} (ID: {request.user.id}) to‘lovni bekor qilishga ruxsatsiz urindi - {timezone.now()}")
        return redirect('/accounting/')
    accounting = Accounting.objects.get(order__order_id=order_id)
    accounting.payment_status = 'jarayonda'
    accounting.updated_by = request.user
    accounting.save()
    logger.info(
        f"Buyurtma #{order_id} to‘lovi bekor qilindi - Bugalter: {request.user.login} (ID: {request.user.id}) - {timezone.now()}")
    messages.success(request, "To‘lov muvaffaqiyatli bekor qilindi!")
    return redirect('/accounting/')


from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q
from django.core.paginator import Paginator
from .models import Accounting
from datetime import datetime
from django.utils import timezone


class AccountingTemplateView(LoginRequiredMixin, TemplateView):
    template_name = 'accounting.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        # Foydalanuvchi ruxsatini tekshirish
        if user.status not in ['bugalter', 'admin', 'filter']:
            context['accounting_entries'] = Accounting.objects.none()
            return context

        # Barcha yozuvlarni olish
        accounting_entries = Accounting.objects.all()

        # Search va status filter
        search_query = self.request.GET.get('search', '')
        if search_query:
            accounting_entries = accounting_entries.filter(
                Q(order__order_id__icontains=search_query) |
                Q(order__client__icontains=search_query) |
                Q(order__filter_data__firma_nomi__icontains=search_query)
            )

        payment_status = self.request.GET.get('status', '')
        if payment_status in ['jarayonda', 'tolandi']:
            accounting_entries = accounting_entries.filter(payment_status=payment_status)

        # Sana bo‘yicha filter
        filter_date = self.request.GET.get('filter_date', '')
        if filter_date:
            try:
                selected_date = datetime.strptime(filter_date, '%Y-%m-%d').date()
                accounting_entries = accounting_entries.filter(
                    order__created_at__date=selected_date
                )
            except ValueError:
                pass

        # Oy bo‘yicha filter
        filter_month = self.request.GET.get('filter_month', '')
        if filter_month:
            try:
                year, month = filter_month.split('-')
                accounting_entries = accounting_entries.filter(
                    order__created_at__year=int(year),
                    order__created_at__month=int(month)
                )
            except ValueError:
                pass

        # Filtrdan keyin nechta yozuv qaytganini tekshirish
        print(f"Sahifada filtrdan keyin: {accounting_entries.count()} yozuv topildi")

        # Pagination
        paginator = Paginator(accounting_entries, 6)  # Har bir sahifada 6 ta yozuv
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Context’ga qo‘shish
        context['accounting_entries'] = page_obj
        context['page_obj'] = page_obj
        context['search_query'] = search_query
        context['payment_status'] = payment_status
        return context


class StatisticsTemplateView(LoginRequiredMixin, TemplateView):
    template_name = 'statistics.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        # Filter turi (kun, oy, yil)
        filter_type = self.request.GET.get('filter_type', 'month')
        filter_date = self.request.GET.get('filter_date')
        if filter_date:
            selected_date = datetime.strptime(filter_date, '%Y-%m-%d')
            if filter_type == 'day':
                start_date = selected_date
                end_date = start_date + timedelta(days=1) - timedelta(seconds=1)
            elif filter_type == 'month':
                start_date = selected_date.replace(day=1)
                end_date = (start_date + timedelta(days=31)).replace(day=1) - timedelta(seconds=1)
            elif filter_type == 'year':
                start_date = selected_date.replace(month=1, day=1)
                end_date = start_date.replace(month=12, day=31) + timedelta(seconds=-1)
        else:
            now = timezone.now()
            start_date = now.replace(day=1)
            end_date = (start_date + timedelta(days=31)).replace(day=1) - timedelta(seconds=1)

        if user.status == 'manager':
            orders = Order.objects.filter(manager=user, created_at__range=[start_date, end_date])
            total_orders = orders.count()
            stats = []
            for order in orders:
                dispecher_price = order.closure.close_price if order.close_order_status else 0
                profit = float(order.price) - float(dispecher_price) if dispecher_price else 0
                stats.append({
                    'order_id': order.order_id,
                    'qayerdan': order.qayerdan,
                    'qayerga': order.qayerga,
                    'manager_price': order.price,
                    'dispecher_price': dispecher_price,
                    'profit': profit,
                    'created_at': order.created_at
                })
            context['stats'] = stats
            context['total_orders'] = total_orders
            context['role'] = 'manager'

        elif user.status == 'dispecher':
            closures = OrderClosure.objects.filter(dispecher=user, closed_at__range=[start_date, end_date]).distinct()
            total_closures = closures.count()
            stats = []
            for closure in closures:
                stats.append({
                    'order_id': closure.order.order_id,
                    'qayerdan': closure.order.qayerdan,
                    'qayerga': closure.order.qayerga,
                    'dispecher_price': closure.close_price,
                    'price_turi': closure.price_turi,
                    'closed_at': closure.closed_at
                })
            context['stats'] = stats
            context['total_orders'] = total_closures
            context['role'] = 'dispecher'

        else:
            context['stats'] = []
            context['total_orders'] = 0
            context['role'] = None

        return context


def mark_paid_view(request, order_id):
    if request.user.status != 'bugalter':
        messages.error(request, "Faqat bugalterlar to‘lovni belgilashi mumkin.")
        logger.warning(
            f"Foydalanuvchi {request.user.login} (ID: {request.user.id}) to‘lovni belgilashga ruxsatsiz urindi - {timezone.now()}")
        return redirect('/accounting/')
    accounting = Accounting.objects.get(order__order_id=order_id)
    accounting.payment_status = 'tolandi'
    accounting.updated_by = request.user
    accounting.save()
    logger.info(
        f"Buyurtma #{order_id} to‘landi deb belgilandi - Bugalter: {request.user.login} (ID: {request.user.id}) - {timezone.now()}")
    messages.success(request, "To‘lov muvaffaqiyatli belgilandi!")
    return redirect('/accounting/')


class OrderListCreateAPIView(generics.ListCreateAPIView):
    serializer_class = OrderSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.status == 'manager':
            return Order.objects.filter(manager=user)
        elif user.status == 'dispecher':
            return Order.objects.all()
        elif user.status == 'admin':
            return Order.objects.all()
        return Order.objects.none()

    def perform_create(self, serializer):
        if self.request.user.status != 'manager':
            raise PermissionDenied("Faqat menejerlar buyurtma yaratishi mumkin.")
        order = serializer.save(manager=self.request.user)
        speak_local("New zayavka")


class OrderDetailAPIView(generics.RetrieveAPIView):
    queryset = Order.objects.all()
    serializer_class = OrderSerializer
    lookup_field = 'order_id'


class OrderClosureAPIView(generics.CreateAPIView):
    queryset = OrderClosure.objects.all()
    serializer_class = OrderClosureSerializer
    permission_classes = [IsDispecher]

    def perform_create(self, serializer):
        order = Order.objects.get(order_id=self.kwargs['order_id'])
        if hasattr(order, 'closure'):
            raise ValidationError("Bu buyurtma allaqachon yopilgan!")
        serializer.save(dispecher=self.request.user, order=order)
        order.close_order_status = True
        order.save()

    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        logger.info(
            f"Buyurtma #{self.kwargs['order_id']} API orqali yopildi - Dispecher: {request.user.login} (ID: {request.user.id}) - {timezone.now()}")
        return Response({"message": "Buyurtma muvaffaqiyatli yopildi va filterchiga yuborildi!"}, status=201)


class AccountingListAPIView(generics.ListAPIView):
    queryset = Accounting.objects.all()
    serializer_class = AccountingSerializer
    permission_classes = [IsBugalter]


class AccountingUpdateAPIView(generics.UpdateAPIView):
    queryset = Accounting.objects.all()
    serializer_class = AccountingSerializer
    permission_classes = [IsBugalter]
    lookup_field = 'order__order_id'

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)
        logger.info(
            f"Buyurtma #{self.kwargs['order__order_id']} hisobi yangilandi - Bugalter: {self.request.user.login} (ID: {request.user.id}) - {timezone.now()}")


@login_required
def firma_suggestions(request):
    query = request.GET.get('q', '').lower()
    if not query:
        return JsonResponse([], safe=False)
    firma_names = FilterData.objects.filter(firma_nomi__istartswith=query).values_list('firma_nomi',
                                                                                       flat=True).distinct()
    return JsonResponse(list(firma_names), safe=False)


@login_required
def manager_stats(request):
    try:
        logger.info(
            f"Foydalanuvchi: {request.user.login} (ID: {request.user.id}), Status: {request.user.status} menejer statistikasini ko‘rmoqchi - {timezone.now()}")
        if request.user.status not in ['admin', 'direktor']:
            logger.warning(
                f"Foydalanuvchi {request.user.login} (ID: {request.user.id}) statistikani ko‘rishga ruxsati yo‘q - {timezone.now()}")
            return JsonResponse({'error': 'Faqat admin va direktorlar statistikani ko‘rishi mumkin.'}, status=403)

        filter_param = request.GET.get('filter', 'all')
        search_param = request.GET.get('search', '').lower()

        managers = Users.objects.filter(status='manager')
        logger.info(f"Topilgan menejerlar soni: {managers.count()} - {timezone.now()}")

        stats = []
        for manager in managers:
            orders = Order.objects.filter(manager=manager)
            if search_param and search_param not in manager.full_name.lower():
                continue

            total_orders = orders.count()
            paid_orders = orders.filter(accounting__payment_status='tolandi').count()
            pending_orders = orders.filter(accounting__payment_status='jarayonda').count()

            if filter_param == 'tolandi' and paid_orders == 0:
                continue
            elif filter_param == 'jarayonda' and pending_orders == 0:
                continue

            stats.append({
                'manager_name': manager.full_name,
                'total_orders': total_orders,
                'paid_orders': paid_orders,
                'pending_orders': pending_orders
            })

        logger.info(f"Menejerlar statistikasi natijalari: {stats} - {timezone.now()}")
        return JsonResponse(stats, safe=False)

    except Exception as e:
        logger.error(
            f"Manager stats endpointida xatolik: {str(e)} - Foydalanuvchi: {request.user.login} (ID: {request.user.id}) - {timezone.now()}")
        return JsonResponse({'error': f"Xatolik yuz berdi: {str(e)}"}, status=500)


@login_required
def dispatcher_stats(request):
    try:
        logger.info(
            f"Foydalanuvchi: {request.user.login} (ID: {request.user.id}), Status: {request.user.status} dispecher statistikasini ko‘rmoqchi - {timezone.now()}")
        if request.user.status not in ['admin', 'direktor']:
            logger.warning(
                f"Foydalanuvchi {request.user.login} (ID: {request.user.id}) statistikani ko‘rishga ruxsati yo‘q - {timezone.now()}")
            return JsonResponse({'error': 'Faqat admin va direktorlar statistikani ko‘rishi mumkin.'}, status=403)

        filter_param = request.GET.get('filter', 'all')
        search_param = request.GET.get('search', '').lower()

        dispatchers = Users.objects.filter(status='dispecher')
        logger.info(f"Topilgan dispecherlar soni: {dispatchers.count()} - {timezone.now()}")

        stats = []
        for dispatcher in dispatchers:
            orders = Order.objects.filter(closure__dispecher=dispatcher)
            if search_param and search_param not in dispatcher.full_name.lower():
                continue

            total_orders = orders.count()
            paid_orders = orders.filter(accounting__payment_status='tolandi').count()
            pending_orders = orders.filter(accounting__payment_status='jarayonda').count()

            if filter_param == 'tolandi' and paid_orders == 0:
                continue
            elif filter_param == 'jarayonda' and pending_orders == 0:
                continue

            stats.append({
                'dispatcher_name': dispatcher.full_name,
                'total_orders': total_orders,
                'paid_orders': paid_orders,
                'pending_orders': pending_orders
            })

        logger.info(f"Dispecherlar statistikasi natijalari: {stats} - {timezone.now()}")
        return JsonResponse(stats, safe=False)

    except Exception as e:
        logger.error(
            f"Dispatcher stats endpointida xatolik: {str(e)} - Foydalanuvchi: {request.user.login} (ID: {request.user.id}) - {timezone.now()}")
        return JsonResponse({'error': f"Xatolik yuz berdi: {str(e)}"}, status=500)


@login_required
def edit_order_page(request, order_id):
    try:
        order = Order.objects.get(order_id=order_id)
        if request.user.status not in ['manager', 'admin', 'direktor']:
            messages.error(request, "Faqat menejerlar, adminlar yoki direktorlar buyurtmani tahrirlashi mumkin.")
            logger.warning(
                f"Foydalanuvchi {request.user.login} (ID: {request.user.id}) buyurtma tahrirlashga ruxsatsiz urindi - {timezone.now()}")
            return redirect('/')
        if request.method == 'POST':
            qayerdan = request.POST.get('qayerdan')
            qayerga = request.POST.get('qayerga')
            client = request.POST.get('client')
            price = request.POST.get('price')
            description = request.POST.get('description')

            order.qayerdan = qayerdan
            order.qayerga = qayerga
            order.client = client
            order.price = price
            order.description = description
            order.save()

            logger.info(
                f"Buyurtma #{order.order_id} tahrirlandi - Foydalanuvchi: {request.user.login} (ID: {request.user.id}) - {timezone.now()}")
            messages.success(request, "Buyurtma muvaffaqiyatli tahrirlandi!")
            return redirect('/')
        return render(request, 'edit_order.html', {'order': order})
    except Order.DoesNotExist:
        messages.error(request, "Buyurtma topilmadi!")
        return redirect('/')


@csrf_exempt
@require_http_methods(["PUT"])
def update_order(request, order_id):
    try:
        order = Order.objects.get(order_id=order_id)
        data = json.loads(request.body)
        order.qayerdan = data.get('qayerdan', order.qayerdan)
        order.qayerga = data.get('qayerga', order.qayerga)
        order.client = data.get('client', order.client)
        order.price = data.get('price', order.price)
        order.description = data.get('description', order.description)
        order.save()
        logger.info(
            f"Buyurtma #{order_id} API orqali tahrirlandi - Foydalanuvchi: {request.user.login} (ID: {request.user.id}) - {timezone.now()}")
        return JsonResponse({'message': 'Buyurtma tahrirlandi!'})
    except Order.DoesNotExist:
        return JsonResponse({'error': 'Buyurtma topilmadi!'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.http import JsonResponse, HttpResponseRedirect
from django.shortcuts import redirect
from django.contrib import messages
from django.utils import timezone
from .models import Order
import logging


@csrf_exempt
@require_http_methods(["DELETE", "POST", "GET"])
def delete_order(request, order_id):
    try:
        order = Order.objects.get(order_id=order_id)
        if order.close_order_status:
            if request.method in ["POST", "GET"]:
                messages.error(request, "Yopilgan buyurtmani o‘chirib bo‘lmaydi!")
                return redirect('/')
            return JsonResponse({'error': 'Yopilgan buyurtmani o‘chirib bo‘lmaydi!'}, status=400)

        # Menejer o‘zi yaratgan buyurtmani o‘chirishi mumkinligini tekshirish
        if request.user.status == 'manager' and order.manager.id != request.user.id:
            if request.method in ["POST", "GET"]:
                messages.error(request, "Bu buyurtmani o‘chirishga ruxsatingiz yo‘q!")
                return redirect('/')
            return JsonResponse({'error': 'Bu buyurtmani o‘chirishga ruxsatingiz yo‘q!'}, status=403)

        order.delete()
        logger.info(
            f"Buyurtma #{order_id} o‘chirildi - Foydalanuvchi: {request.user.login} (ID: {request.user.id}) - {timezone.now()}")

        if request.method in ["POST", "GET"]:
            messages.success(request, "Buyurtma muvaffaqiyatli o‘chirildi!")
            return redirect('/')
        return JsonResponse({'message': 'Buyurtma o‘chirildi!'})

    except Order.DoesNotExist:
        logger.warning(
            f"Buyurtma #{order_id} topilmadi - Foydalanuvchi: {request.user.login} (ID: {request.user.id}) - {timezone.now()}")
        if request.method in ["POST", "GET"]:
            messages.error(request, "Buyurtma topilmadi!")
            return redirect('/')
        return JsonResponse({'error': 'Buyurtma topilmadi!'}, status=404)
    except Exception as e:
        logger.error(f"Buyurtma o‘chirishda xato: {str(e)} - Order ID: {order_id} - {timezone.now()}")
        if request.method in ["POST", "GET"]:
            messages.error(request, f"Xatolik yuz berdi: {str(e)}")
            return redirect('/')
        return JsonResponse({'error': str(e)}, status=400)


class ManagerStatsTemplateView(LoginRequiredMixin, TemplateView):
    template_name = 'manager_stats.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.status not in ['admin', 'direktor']:
            messages.error(self.request, "Faqat admin va direktorlar statistikani ko‘rishi mumkin.")
            logger.warning(
                f"Foydalanuvchi {self.request.user.login} (ID: {self.request.user.id}) menejer statistikasini ko‘rishga ruxsatsiz urindi - {timezone.now()}")
        return context


class DispatcherStatsTemplateView(LoginRequiredMixin, TemplateView):
    template_name = 'dispatcher_stats.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.user.status not in ['admin', 'direktor']:
            messages.error(self.request, "Faqat admin va direktorlar statistikani ko‘rishi mumkin.")
            logger.warning(
                f"Foydalanuvchi {self.request.user.login} (ID: {self.request.user.id}) dispecher statistikasini ko‘rishga ruxsatsiz urindi - {timezone.now()}")
        return context


class NotificationViewList(LoginRequiredMixin, ListView):
    template_name = 'notifications.html'
    context_object_name = 'notifications'

    def get_queryset(self):
        return Notification.objects.filter(recipient=self.request.user).order_by('-timestamp')


@require_http_methods(["POST"])
def toggle_theme(request):
    current_theme = request.session.get('theme', 'light')
    new_theme = 'dark' if current_theme == 'light' else 'light'
    request.session['theme'] = new_theme
    return redirect(request.META.get('HTTP_REFERER', '/'))


from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.contrib import messages
from django.db.models import Q
from django.utils import timezone
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .models import Accounting


@login_required
def export_accounting_excel(request):
    if request.user.status not in ['bugalter', 'admin']:
        messages.error(request, "Faqat bugalterlar va adminlar eksport qilishi mumkin.")
        return redirect('/accounting/')

    # GET parametrlarni tekshirish
    print(f"Excel eksport uchun GET parametrlari: {request.GET}")

    # Accounting ma'lumotlarini olish
    accounting_entries = Accounting.objects.all()

    # Search va status filter
    search_query = request.GET.get('search', '')
    if search_query:
        accounting_entries = accounting_entries.filter(
            Q(order__order_id__icontains=search_query) |
            Q(order__client__icontains=search_query) |
            Q(order__filter_data__firma_nomi__icontains=search_query)
        )

    payment_status = request.GET.get('status', '')
    if payment_status in ['jarayonda', 'tolandi']:
        accounting_entries = accounting_entries.filter(payment_status=payment_status)

    # Sana bo‘yicha filter
    filter_date = request.GET.get('filter_date', '')
    if filter_date:
        try:
            selected_date = datetime.strptime(filter_date, '%Y-%m-%d').date()
            accounting_entries = accounting_entries.filter(
                order__created_at__date=selected_date
            )
        except ValueError:
            pass

    # Oy bo‘yicha filter
    filter_month = request.GET.get('filter_month', '')
    if filter_month:
        try:
            year, month = filter_month.split('-')
            accounting_entries = accounting_entries.filter(
                order__created_at__year=int(year),
                order__created_at__month=int(month)
            )
        except ValueError:
            pass

    # Filtrdan keyin nechta yozuv qaytganini tekshirish
    print(f"Excel eksport uchun filtrdan keyin: {accounting_entries.count()} yozuv topildi")

    # Excel faylini yaratish
    wb = Workbook()
    ws = wb.active
    ws.title = "Accounting"

    headers = ['Order ID', 'Client', 'Manager Price', 'Dispecher Price', 'Payment Type', 'Payment Status', 'Updated At']
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))

    for row, entry in enumerate(accounting_entries, start=2):
        ws.append([
            entry.order.order_id,
            entry.order.client,
            float(entry.manager_price),
            float(entry.dispecher_price),
            entry.payment_type,
            entry.payment_status,
            entry.order.created_at.strftime('%Y-%m-%d %H:%M')
        ])
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="accounting_report.xlsx"'
    wb.save(response)
    return response


from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.contrib import messages
from django.db.models import Q
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import io
from .models import Accounting


@login_required
def export_accounting_pdf(request):
    if request.user.status not in ['bugalter', 'admin']:
        messages.error(request, "Faqat bugalterlar va adminlar eksport qilishi mumkin.")
        return redirect('/accounting/')

    # GET parametrlarni tekshirish
    print(f"PDF eksport uchun GET parametrlari: {request.GET}")

    # Accounting ma'lumotlarini olish
    accounting_entries = Accounting.objects.all()

    # Search va status filter
    search_query = request.GET.get('search', '')
    if search_query:
        accounting_entries = accounting_entries.filter(
            Q(order__order_id__icontains=search_query) |
            Q(order__client__icontains=search_query) |
            Q(order__filter_data__firma_nomi__icontains=search_query)
        )

    payment_status = request.GET.get('status', '')
    if payment_status in ['jarayonda', 'tolandi']:
        accounting_entries = accounting_entries.filter(payment_status=payment_status)

    # Sana bo‘yicha filter
    filter_date = request.GET.get('filter_date', '')
    if filter_date:
        try:
            selected_date = datetime.strptime(filter_date, '%Y-%m-%d').date()
            accounting_entries = accounting_entries.filter(
                order__created_at__date=selected_date
            )
        except ValueError:
            pass

    # Oy bo‘yicha filter
    filter_month = request.GET.get('filter_month', '')
    if filter_month:
        try:
            year, month = filter_month.split('-')
            accounting_entries = accounting_entries.filter(
                order__created_at__year=int(year),
                order__created_at__month=int(month)
            )
        except ValueError:
            pass

    # Filtrdan keyin nechta yozuv qaytganini tekshirish
    print(f"PDF eksport uchun filtrdan keyin: {accounting_entries.count()} yozuv topildi")

    # PDF faylini yaratish
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    styles = getSampleStyleSheet()
    title = Paragraph("Accounting Report", styles['Title'])
    elements.append(title)

    data = [['Order ID', 'Client', 'Manager Price', 'Dispecher Price', 'Payment Type', 'Payment Status', 'Updated At']]
    for entry in accounting_entries:
        data.append([
            entry.order.order_id,
            entry.order.client,
            f"{entry.manager_price:.2f}",
            f"{entry.dispecher_price:.2f}",
            entry.payment_type,
            entry.payment_status,
            entry.order.created_at.strftime('%Y-%m-%d %H:%M')
        ])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="accounting_report.pdf"'
    return response


class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        context['total_orders'] = Order.objects.count()
        context['open_orders'] = Order.objects.filter(close_order_status=False).count()
        context['closed_orders'] = Order.objects.filter(close_order_status=True).count()
        context['recent_orders'] = Order.objects.order_by('-created_at')[:5]
        return context


class ChatTemplateView(TemplateView):
    template_name = 'chat.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        messages = ChatMessage.objects.all().order_by('created_at')
        context['messages'] = messages
        return context


def post_chat_message(request):
    if request.method == 'POST':
        content = request.POST.get('content')
        user_name = request.POST.get('user_name')  # Foydalanuvchi nomi formadan olinadi

        if not user_name:
            user_name = "Anonim"  # Agar ism kiritilmagan bo‘lsa, "Anonim" deb belgilaymiz

        if content:
            # Agar foydalanuvchi tizimga kirgan bo‘lsa, user ni saqlaymiz, aks holda None
            user = request.user if request.user.is_authenticated else None
            ChatMessage.objects.create(
                user=user,
                user_name=user_name,  # Foydalanuvchi nomi alohida saqlanadi
                content=content
            )
            logger.info(
                f"Yangi xabar yuborildi - Foydalanuvchi: {user_name} - {timezone.now()}"
            )
            messages.success(request, "Xabar muvaffaqiyatli yuborildi!")
        else:
            messages.error(request, "Xabar bo‘sh bo‘lmasligi kerak!")
    return redirect('/chat/')
